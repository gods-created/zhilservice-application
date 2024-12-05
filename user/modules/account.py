from copy import deepcopy
from boto3 import resource
from botocore.exceptions import ClientError
from os import getenv, remove, path
from openpyxl import load_workbook
from xlrd import open_workbook
from re import search
from ..tasks import send_message_to_mail

class Account:
    def __init__(self):
        self.response_json = {
            'status': 'error',
            'err_description': ''
        }

        self.pattern = rf'лс:\s*[NUMBER]\s*СС:'
        self.s3_bucket = None

    async def __aenter__(self):
        aws_access_key_id = getenv('AWS_ACCESS_KEY_ID')
        aws_secret_access_key = getenv('AWS_SECRET_ACCESS_KEY')
        aws_bucket_name = getenv('AWS_BUCKET_NAME')

        if all((
            aws_access_key_id, aws_secret_access_key, aws_bucket_name
        )):
            self.s3_bucket = resource(
                's3',
                aws_access_key_id=aws_access_key_id,
                aws_secret_access_key=aws_secret_access_key,
                region_name='us-east-1'
            ).Bucket(aws_bucket_name)

        return self
    
    def __download_file(self, locality: str) -> dict:
        response_json = deepcopy(self.response_json)

        try:
            s3_bucket = self.s3_bucket
            if not s3_bucket:
                raise Exception('Неможливо підключитися до ресурсу для завантежння файлу з рахунками.')
            
            filename = ''
            filenames = ('zarichne.xlsx', 'zarichne.xls', ) if locality == 'Зарічне' else ('cherkaske.xlsx', 'cherkaske.xls', )
            
            for i in filenames:
                try:
                    s3_bucket.download_file(i, i)
                    filename = i
                    break
                except (ClientError, ) as e:
                    continue

            if not filename:
                raise Exception(f'Файл з рахунками не знайдений.')

            response_json['status'] = 'success'
            response_json['filename'] = filename
            
        except (Exception, ) as e:
            response_json['err_description'] = str(e)

        finally:
            return response_json
        
    def __translit_process(self, string: str) -> str:
        if not isinstance(string, str):
            return ''
        
        transliteration_map = {
            'À': 'А', 'Á': 'Б', 'Â': 'В', 'Ã': 'Г', 'Ä': 'Д', 'Å': 'Е', 'Æ': 'Ж',
            'Ç': 'З', 'È': 'И', 'É': 'Й', 'Ê': 'К', 'Ë': 'Л', 'Ì': 'М', 'Í': 'Н',
            'Î': 'О', 'Ï': 'П', 'Ð': 'Р', 'Ñ': 'С', 'Ò': 'Т', 'Ó': 'У', 'Ô': 'Ф',
            'Õ': 'Х', 'Ö': 'Ц', '×': 'Ч', 'Ø': 'Ш', 'Ù': 'Щ', 'Ú': 'Ъ', 'Û': 'Ы',
            'Ü': 'Ь', 'Ý': 'Э', 'Þ': 'Ю', 'ß': 'Я', 
            'à': 'а', 'á': 'б', 'â': 'в', 'ã': 'г', 'ä': 'д', 'å': 'е', 'æ': 'ж',
            'ç': 'з', 'è': 'и', 'é': 'й', 'ê': 'к', 'ë': 'л', 'ì': 'м', 'í': 'н',
            'î': 'о', 'ï': 'п', 'ð': 'р', 'ñ': 'с', 'ò': 'т', 'ó': 'у', 'ô': 'ф',
            'õ': 'х', 'ö': 'ц', '÷': 'ч', 'ø': 'ш', 'ù': 'щ', 'ú': 'ъ', 'û': 'ы',
            'ü': 'ь', 'ý': 'э', 'þ': 'ю', 'ÿ': 'я',
            'Ë': 'Ё', 'ё': 'ё'
        }

        result = []
        for char in string:
            result.append(transliteration_map.get(char, char))

        converted_text = ''.join(result)
        return converted_text
        
    def __xlsx(self, current_account_number: str, filename: str) -> list:
        data = []

        workbook = load_workbook(filename, data_only=True)
        sheet_obj = workbook.active
        rows = sheet_obj.max_row
        cols = sheet_obj.max_column
        pattern = self.pattern.replace('[NUMBER]', current_account_number)

        for i in range(1, rows):
            cell = sheet_obj.cell(row=i, column=1)
            cell_value = self.__translit_process(
                cell.value
            )
            if search(pattern, cell_value):
                for j in range(2, cols):
                    cell_obj = sheet_obj.cell(row=i, column=j)
                    data.append(
                        cell_obj.value
                    )
                
                break

        return data 
    
    def __xls(self, current_account_number: str, filename: str) -> list:
        data = []

        workbook = open_workbook(filename)
        sheet = workbook.sheet_by_index(0)
        rows = sheet.nrows
        cols = sheet.ncols
        pattern = self.pattern.replace('[NUMBER]', current_account_number)

        for i in range(1, rows):
            cell_value = self.__translit_process(
                sheet.cell_value(i, 1)
            )

            if search(pattern, cell_value):
                for j in range(2, cols):
                    data.append(sheet.cell_value(i, j))
                break

        return data 
        
    def __get_account_info(self, current_account_number: str, filename: str) -> dict:
        response_json = deepcopy(self.response_json)

        try:
            if not path.exists(filename) or not filename.endswith(('.xlsx', '.xls')):
                raise FileNotFoundError('Помилка доступу до файлу з рахунками.')
            
            data = self.__xlsx(current_account_number, filename) if filename.endswith('.xlsx') else self.__xls(current_account_number, filename)
            remove(filename)
            
            response_json['data'] = data
            response_json['status'] = 'success'

        except (FileNotFoundError, Exception, ) as e:
            response_json['err_description'] = str(e)

        finally:
            return response_json
        
    def __send_message_to_mail(self, current_account_number: str, email: str, data: list) -> dict:
        response_json = deepcopy(self.response_json)

        try:
            send_message_to_mail.apply_async(
                (current_account_number, email, data),
                queue='high_priority'
            )

            response_json['status'] = 'success'

        except (Exception, ) as e:
            response_json['err_description'] = str(e)

        finally:
            return response_json
    
    async def get(self, data: dict) -> dict:
        response_json = deepcopy(self.response_json)

        try:
            email, current_account_number, locality = data.get('email'), data.get('current_account_number'), data.get('locality')
            if not all((
                email, current_account_number, locality
            )):
                raise Exception('Некоректно заповнена форма.')

            download_file_response = self.__download_file(locality)
            if download_file_response['status'] == 'error':
                raise Exception(
                    download_file_response['err_description']
                )
            
            filename = download_file_response.get('filename', '')
            get_account_info_response = self.__get_account_info(current_account_number, filename)
            if get_account_info_response['status'] == 'error':
                raise Exception(
                    get_account_info_response['err_description']
                )

            data = get_account_info_response.get('data', [])
            send_message_to_mail_response = self.__send_message_to_mail(current_account_number, email, data)
            if send_message_to_mail_response['status'] == 'error':
                raise Exception(
                    send_message_to_mail_response['err_description']
                )
            
            # response_json['data'] = data
            response_json['status'] = 'success'

        except (Exception, ) as e:
            response_json['err_description'] = str(e)

        finally:
            return response_json
    
    async def __aexit__(self, *args, **kwargs):
        pass